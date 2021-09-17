#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#
# Para uso único de los proyectos de Solidaridad en Marcha o
# procesos de aprendizaje autorizados por el mismo ente
#
# Creador: Rodolfo Piedra Camacho
# Contacto: fofo.piedra@gmail.com
#

"""
Script para distribución de Censos con sus Donantes Respectivos

El script cumple las siguientes funciones
 - Traduce una encuesta en formato tipo .csv a formato .xlsx
 - Distribuye las solicitudes de la encuesta con los respectivos censos
 - Reconoce casos de error y los documenta en un tab llamado "Crisis". Errores:
   - No exiten censos que distribuir

Detalles Importantes y Cuidados
###############################
La sección "VARIABLES GLOBALES" es crítica para el funcionamiento del script
modificaciones en los excels a utilizar normalmente requieren un cambio
respectivo en esta sección. Esto es especialmente sensible en archivos que usan
fechas o lugares como parte de su nombre, ejemplos de variables globales:
 - Nombres de archivos .xlsx y .csv
 - Número/Letra de Columna donde se almacenan los datos respectivos (Nombres
   o telefeonos en las encuestas)

Informacion de las
Cada hoja con información de los censos debe respetar el siguiente formato:
 1) Columna A debe estar vacia y solo poseera una x para marcar la
    separación entre un censo y otro
 2) Solo debe de haber una fila de separación entre censo y censo. En la
    columna A de esta fila se debe digitar una "x". La primera fila del
    archivo respeta esta convención.
 NOTA: El digito de separacion esperado "x" y la columna vacia esperada "A" se
       pueden modificar con las variables globales "separador_censos" y
       "columna_dueño" respectivamente
 3) El último censo en una hoja debe ser seguido por el símbolo "yy"

Debe existir una hoja con el mismo nombre que la variable global
hoja_data_delicada. En esta hoja en las Celdas marcadas por lista_celdas
debe de encontrarse un número que indica la última fila del archivo
de encuestas analizadas. En un primer uso del script debe escribirse a 0.

La hoja con la información de los dueños de censo debe respetar el siguiente
formato:
 1) La primera fila no debe poseer información valiosa, solo los headers
    de la tabla

LIMITACIONES
###############################
1) La información en la hoja de encuestas no es corroborada de ninguna forma
2) No existe método para borrar entradas
3) Las variables globales solo se pueden modificar previo a usar el script por
   primera vez en una familia de archivos. Si se modifican los valores y se
   reusa sobre un archivo ya trabajado los resultados son desconocidos
"""
###############################################################################
#                                  IMPORTS                                    #
###############################################################################
import csv
import openpyxl
import logging
from copy import deepcopy

###############################################################################
#                                  GLOBAL                                    #
###############################################################################
# Loggers
# ------------------------------------------------------------------------------
# Variable de Log para documentar errores de ejecución
log = logging.getLogger(__name__)

# Filenames
# ------------------------------------------------------------------------------
# Nombre del archivo donde se deja la información final
resultados_filename = "Entregas NEJ 2018.xlsx"

# Nombre original del archivo de encuesta, extensión .csv
csv_orig_filename = "Cajas de la Amor.csv"

# Nombre del archivo de encuestas post conversión csv->xlsx
encuesta_filename = "encuesta.xlsx"

# Nombre de la hoja en el archivo de encuestas post conversión csv->xlsx
encuesta_sheet_name = "DATA"

# Nombre del archivo con información de los censos
censos_filename = "Entregas NEJ 2018.xlsx"

# Nombres de las Hojas con Información de Censos
lugar_censo_1 = "DUL"
lugar_censo_2 = "PUR"
lugar_censo_3 = "LIA"
lugar_censo_4 = "QUI"
lugar_censo_5 = "CPN"
lugar_censo_6 = "CRP"

# Nombres de la Hoja donde se almacena la información de Error
hoja_error = "CRISIS"

# Nombres de la Hoja donde se almacena la información de Delicada
hoja_data_delicada = "DATA_DELICADA"

# Lista de Celdas donde Se Almacena el Contador de Distribución
lista_celdas = ["A1", "A10", "A20", "A100"]

# Informacion Sobre Centros de Acopio y su Relacion con los Censos
# ------------------------------------------------------------------------------
# Diccionario de diccionarios con la informacion de los centros de acopio
# Primer Nivel de Diccionario: Usa como llaves los nombres de los centros de
#                              acopios estas llaves retornan otro diccionario
# Segundo Nivel de Diccionario: Usa como llaves los nombres de los censos
#                               (variables de nombre "lugar_censo_#"). Esto
#                               asocia el centro de acopio a una censo/fiesta.
#                               Estas llaves retornan una lista vacia
# NOTA: La lista debe estar vacia. Es de uso exclusivo del script
dict_centros_de_acopio = {"Tres Rios/Curridabat": {lugar_censo_1: []},
                          "Santa Ana":            {lugar_censo_3: [],
                                                   lugar_censo_4: [],
                                                   lugar_censo_5: []},
                          "Escazu":               {lugar_censo_2: [],
                                                   lugar_censo_6: []}
                          }

# Diccionario que mantiene el status de los distintos centros de acopio
dict_centros_de_acopio_full_status = {"Tres Rios/Curridabat": False,
                                      "Santa Ana":            False,
                                      "Escazu":               False}

# Separador entre censos
# El caracter debe aparecer entre cada censo, esto permite al script detectar
# el inicio y fin de un censo
separador_censos = "x"

# Break de lectura de censos
# Si se encuentra este caracter la funcion censos_loader parara de sacar
# informacion de una hoja de censos dada, ignorando todos los censos que puedan
# aparecer después.
# Este caracter no es encesario para la lectura completa de los censos y debe
# de ser usado por terminos de debug
break_lectura_censos = "yy"

# Columna donde se espera encontrar el separador de censos y donde se
# escribira el nombre y numero de teléfono del dueño
columna_dueño = "A"

# Localización de Datos en .csv y .xlsx de Encuesta
# A su vez se usa como las columnas donde se almacena la información de Error
# ------------------------------------------------------------------------------
# Columna donde se espera encontrar el nombre del dueño de un censo
columna_nombre = "B"

# Columna donde se espera encontrar el número de telefono del dueño de un censo
columna_telefono = "C"

# Columna donde se espera encontrar la cantidad de cajas a entregar
columna_cant_cajas = "D"

# Columna donde se espera encontrar la opción preferida de centro de acopio
columna_c_acopio_1 = "E"

# Columna donde se espera encontrar la segunda opción de centro de acopio
columna_c_acopio_2 = "F"

# Columna Código de Error, no puede ser igual a columna_nombre,
# columna_telefono, columna_c_acopio_1 o columna_c_acopio_2
columna_error = "G"

# Códigos de Error
# ------------------------------------------------------------------------------
# Mismo Centro de Acopio
error_repetición = 0

# Fallo en distribución
error_distribución = 1

# Mensajes de Error. Lista indexada por los codigo de error
msj_error = ["No se encontro espacio en el Primer Sector de Distribución y el"
             " Segundo era repetido",
             "No se logro asignar un censo al interesado porque ambos sectores"
             " estan completos"]


###############################################################################
#                                   CLASES                                    #
###############################################################################
class CensoOwner():
    """
    Contiene la información de un encargado de un único censo
    :param nombre:             Nombre del Dueño
    :param telefono:           Número de Teléfono del Dueño
    :param c_acopio_1:         Nombre de la primera opción para centro de
                               acopio
    :param c_acopio_2:         Nombre de la segunda opción para centro de
                               acopio
    :param c_acopio_repetido:  Booleano que marca si el centro de acopio esta
                               repetido
    :param codigo_error:       Almacena el número de error en caso de ser
                               necesario
    """
    def __init__(self, nombre, telefono, c_ac_1, c_ac_2, cant_cajas):
        self.nombre = nombre
        self.telefono = telefono
        self.c_acopio_1 = c_ac_1
        self.c_acopio_2 = c_ac_2
        self.c_acopio_repetido = (c_ac_1 == c_ac_2)
        self.codigo_error = None
        self.cant_cajas = cant_cajas

    # Método para propositos de debugging
    def print_data(self):
        print("\nNombre: {}, Tel: {}, C_Acopio_1: {}, C_Acopio_2 {}, "
              "C_Acopio_Rep {}, Código_error {}".format(
                self.nombre, self.telefono, self.c_acopio_1, self.c_acopio_2,
                self.c_acopio_repetido, self.codigo_error))


###############################################################################
#                                   UTILS                                     #
###############################################################################
# Conversor de CSV a XLSX
# ------------------------------------------------------------------------------
def conv_to_xlsx(arch_csv):
    """
    Carga un archivo de encuestas .csv y lo convierte a formato .xlsx para
    su procesamiento con openpyxl

    :param arch_csv: Nombre del archivo de encuestas con extención .csv
    """
    # Abrir el archivo
    try:
        encuesta_csv = open(arch_csv)

    except Exception:
        log.error("Archivo de encuentas {} no se encuentra".format(arch_csv))
        raise

    # Demarca el signo que usa el .csv para separar su información
    csv.register_dialect('colons', delimiter=',')

    # Carga el contenido del archivo .csv
    lectura = csv.reader(encuesta_csv, dialect='colons')

    # Definir objeto de xlsx con una hoja de nombre específico
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    ws.title = encuesta_sheet_name

    # Conversión del .csv al objeto woorkbook
    for row_index, row in enumerate(lectura):
        for column_index, cell in enumerate(row):
            # Coordenadas de cell empiezan en 1,1
            ws.cell(column=column_index + 1, row=(row_index + 1)).value = cell

    # Guarda el archivo de encuestas
    wb.save(filename=encuesta_filename)


# Parser de Excel de Encuestas
# ------------------------------------------------------------------------------
def data_loader(fila_previa):
    """
    Procesa y almacena la información del archivo de encuesta extensión .xlsx
    en un array de clases tipo CensoOwner

    :param fila_previa: Contiene el valor de la última fila que fue analizada

    :return: Un array de clases tipo CensoOwner
    """
    parsed_data = []

    # Abrir el Excel de encuestas
    wb = openpyxl.load_workbook(encuesta_filename)
    data_sheet = wb.active

    # Ciclo for para iterar sobre todas las filas de la encuesta
    print(data_sheet.max_row)
    print(fila_previa)
    for row in range(fila_previa, data_sheet.max_row):

        # En los objetos de openpyxl no existe el indice 0 para filas
        # o columnas, por eso el numero de fila se debe actualizar
        file_row = row + 1

        # La primera fila contiene los headers del .csv
        if file_row is 1:
            continue

        # Un mismo donador podria entregar más de una caja
        cajas_totales = \
            int(data_sheet[columna_cant_cajas + str(file_row)].value)

        for cant_cajas in range(cajas_totales):
            parsed_data.append(CensoOwner(
                nombre=data_sheet[columna_nombre+str(file_row)].value,
                telefono=data_sheet[columna_telefono+str(file_row)].value,
                c_ac_1=data_sheet[columna_c_acopio_1+str(file_row)].value,
                c_ac_2=data_sheet[columna_c_acopio_2+str(file_row)].value,
                cant_cajas=data_sheet[columna_cant_cajas+str(file_row)].value))

    return parsed_data


# Parser de Excel de Censos
# ------------------------------------------------------------------------------
def censos_loader():
    """
    Procesa y almacena la información del archivo de censos extensión .xlsx
    Guarda en las listas del diccionario de centros de acopios el numero
    de las celdas respectivas a una caja donde no se a asignado a un dueño

    :return: Una copia de la variable global dict_centros_de_acopio con la
             lista de censos sin dueño asociada a cada Censo
    """
    global dict_centros_de_acopio

    # Copia completa al diccionario de centros de acopio
    dict_copy = deepcopy(dict_centros_de_acopio)

    # Abrir archivo con información de censos
    try:
        wb = openpyxl.load_workbook(censos_filename)
    except Exception:
        log.error("No existe el archivo {} en el direcctorio del "
                  "Script".format(censos_filename))
        raise

    # Sanity Check: Asegurar que las hojas son parte del diccionario
    for centro_de_acopio in dict_copy.values():
        for censo in centro_de_acopio.keys():
            if censo not in wb.sheetnames:
                log.error("Censo de nombre {} no es una hoja existente del "
                          "archivo {}".format(censo, censos_filename))
                raise KeyError("Worksheet {} does not exist.".format(censo))

    # Parser de la información
    # Ciclo for para cada centro de acopio en el diccionario principal
    for centro_de_acopio in dict_copy.keys():
        log.debug(centro_de_acopio)

        # Ciclo for para cada Censo dentro del Centro de Acopio Actual
        for censo in dict_copy[centro_de_acopio].keys():
            log.debug(censo)
            hoja_actual = wb[censo]
            log.debug(hoja_actual.max_row)

            # For para analizar cada celda
            for i in range(hoja_actual.max_row):

                # En los objetos de openpyxl no existe el indice 0 para filas
                # o columnas, por eso el numero de fila se debe actualizar
                row = i + 1

                # Deteccion de censo
                # Se sabe que se encontro un censo cuando se encuentra una
                # casilla que posee el símbolo separador_censos.
                # A su vez se sabe que este no tiene un dueño si las siguientes
                # dos filas estan vacias. (Aqui es donde se encontraría la
                # información de dueño y numero de celular)
                censo_libre = (
                    hoja_actual[columna_dueño + str(row)].value ==
                    separador_censos and
                    hoja_actual[columna_dueño + str(row+1)].value is None and
                    hoja_actual[columna_dueño + str(row+2)].value is None)

                if censo_libre:
                    dict_copy[centro_de_acopio][censo].append(row)

                # Detener Lectura
                if hoja_actual[columna_dueño + str(row)].value == \
                   break_lectura_censos:
                    log.info("Se encontro simbolo de break {}, todos los"
                             " censos de {} posteriores a la linea {} no seran"
                             " distribuidos".format(break_lectura_censos,
                                                    censo, row))
                    break

    return dict_copy


# Parser de Excel de Censos
# ------------------------------------------------------------------------------
def censo_spread(data_encuesta, data_censos):
    """
    Algoritmo principal de distribución de cajas. Asocia a cada posible dueño
    con un censo. Modifica el archivo de censos extensión .xlsx con el nombre
    y número de teléfono del encargado

    :param data_encuesta: Lista de objetos CensoOwner con informacion de los
                          posibles dueños de censos
    :param data_censos:   Diccionario de diccionarios con informacion de los
                          centros de acopio y sus censos asociados
    """

    # Abrir archivo con información de censos
    # ---------------------------------------
    try:
        wb = openpyxl.load_workbook(censos_filename)
    except Exception:
        log.error("No existe el archivo {} en el direcctorio del "
                  "Script".format(censos_filename))
        raise

    # Lista de elementos CensoOwner que presentaron error
    lista_error = []

    # Ciclo For principal
    # Recorre toda la lista de info en la encuesta y distribuye cada censo
    # --------------------------------------------------------------------
    for dueño in data_encuesta:
        # dueño.print_data()

        # Booleano de control si el dueño fue asignado o no una caja
        caja_asignada = False

        # Obtenga los posibles centros de acopio deseado
        lista_c_acopio = [dueño.c_acopio_1, dueño.c_acopio_2]

        # For de Centros de Acopio
        for i, c_acopio in enumerate(lista_c_acopio):

            log.debug("#####{}#####".format(c_acopio))
            # For de las Fiestas asignadas al Centro de Acopio
            for censo in data_censos[c_acopio].keys():

                log.debug("----{}----".format(censo))
                cant_familias = len(data_censos[c_acopio][censo])

                # Si no quedan familias en el censo seguir al siguiente
                if cant_familias == 0:
                    log.debug("No more Boxes")
                    continue
                # Si quedan, asignar la primera en al lista al dueño
                else:
                    # Fila a escribir
                    # Se elimina el número a asignar de la lista del censo
                    row = data_censos[c_acopio][censo].pop(0)
                    log.debug("Caja fila: {}, de {}, fue asignada a {}".format(
                        row, censo, dueño.nombre))

                    # Escribe la información del nuevo Dueño
                    wb[censo][columna_dueño+str(row+1)] = dueño.nombre
                    wb[censo][columna_dueño+str(row+2)] = dueño.telefono

                    # Booleano de caja asignada a True
                    caja_asignada = True

                    # La caja fue asignada se puede salir del ciclo For
                    break

            # Si la caja fue asignada se puede seguir con el siguiente dueño
            if caja_asignada:

                # Impresion de Mensaje de Usuario en caso de que se usara
                # la segunda opción de Sector de Entrega
                if i == 1:
                    log.info("Para el dueño de nombre {} y telefono {} se tuvo"
                             "que elegir su segunda opción de sector de "
                             "entrega: {}".format(dueño.nombre,
                                                  dueño.telefono,
                                                  dueño.c_acopio_2))

                break

            # En caso de que un Sector de Entrega este completo, se le presenta
            # el mensaje al usuario
            if not dict_centros_de_acopio_full_status[c_acopio]:
                dict_centros_de_acopio_full_status[c_acopio] = True
                log.info("Sector: {} ha entregado todas sus cajas".format(
                    c_acopio))

            # Safety Check
            if dueño.c_acopio_repetido:
                # CRISIS 1
                log.error("Crisis 1: Para el dueño {} telefono {}".format(
                    dueño.nombre, dueño.telefono))
                dueño.codigo_error = error_repetición
                lista_error.append(dueño)
                break

        # Chequea si el dueño posee una caja asignada
        if not caja_asignada and dueño.codigo_error is None:
            log.error("Crisis 2: Para el dueño {} telefono {}".format(
                dueño.nombre, dueño.telefono))
            dueño.codigo_error = error_distribución
            lista_error.append(dueño)

    # Posterior a la Distribución
    # ---------------------------
    # Resolución de casos de error
    errores_existentes = wb[hoja_error].max_row
    log.debug("Errores Existentes: {}".format(errores_existentes))
    for i, error in enumerate(lista_error):
        # error.print_data()

        # Se escribe la información de Error en la Hoja Crisis
        wb[hoja_error][columna_nombre+str(errores_existentes+i+1)] = \
            error.nombre
        wb[hoja_error][columna_telefono+str(errores_existentes+i+1)] = \
            error.telefono
        wb[hoja_error][columna_c_acopio_1+str(errores_existentes+i+1)] = \
            error.c_acopio_1
        wb[hoja_error][columna_c_acopio_2+str(errores_existentes+i+1)] = \
            error.c_acopio_2
        wb[hoja_error][columna_cant_cajas+str(errores_existentes+i+1)] = \
            error.cant_cajas
        wb[hoja_error][columna_error+str(errores_existentes+i+1)] = \
            msj_error[error.codigo_error]

    # Log del estado de los censos post distirbución
    for centro_de_acopio in data_censos.keys():

        # Ciclo for para cada Censo dentro del Centro de Acopio Actual
        for censo in data_censos[centro_de_acopio].keys():
            log.info("Censo {} de Sector {} posee {} sin distribuir".format(
                censo,
                centro_de_acopio,
                len(data_censos[centro_de_acopio][censo])))

    # Nota de la última fila de encuesta
    wb_encuesta = openpyxl.load_workbook(encuesta_filename)
    data_sheet = wb_encuesta.active
    for celda in lista_celdas:
        wb[hoja_data_delicada][celda].value = data_sheet.max_row

    # Escritura final del archivo
    # ---------------------------
    wb.save(resultados_filename)


###############################################################################
#                                   SETUP                                     #
###############################################################################
def setup():
    # Log Format Setup
    # --------------------------------------------------------------------------
    try:
        from colorlog import ColoredFormatter as Formatter
        logfrmt = (
            '  {thin_white}{asctime}{reset} | '
            '{log_color}{levelname:8}{reset} | '
            '{thin_white}{processName}{reset} | '
            '{log_color}{message}{reset}'
        )
    except ImportError as e:
        from logging import Formatter
        logfrmt = (
            '  {asctime} | '
            '{levelname:8} | '
            '{processName} | '
            '{message}'
        )

    # Set log format
    stream = logging.StreamHandler()
    stream.setFormatter(Formatter(fmt=logfrmt, style='{'))

    # Set logging Level
    level = logging.INFO
    logging.basicConfig(handlers=[stream], level=level)

    log.info('Verbosity at level {}'.format(level))


###############################################################################
#                                MAIN SCRIPT                                  #
###############################################################################
def main():
    # Convertir archivo .csv a .xlsx
    conv_to_xlsx(csv_orig_filename)

    # Sanity Check
    # Se obtiene cantidad de lineas de encuesta previamente leidas
    try:
        wb = openpyxl.load_workbook(censos_filename)
    except Exception:
        log.error("No existe el archivo {} en el direcctorio del "
                  "Script".format(censos_filename))
        raise

    ultima_fila = wb[hoja_data_delicada][lista_celdas[0]].value
    log.info("Ultima Fila distribuida: {}".format(ultima_fila))
    for celda in lista_celdas:
        if ultima_fila != wb[hoja_data_delicada][celda].value:
            log.critical("CORRUPCION EN HOJA {}".format(hoja_data_delicada))
            return

    # Parsear datos de encuesta
    parsed_data = data_loader(ultima_fila)
    """
    for member in parsed_data:
        member.print_data()
    """
    log.debug("Parsed data length: ", len(parsed_data))

    # Asegurar que hay nuevos dueños que asignar
    if len(parsed_data) == 0:
        log.info("No hay nuevos dueños que asignar")
        return

    # Parsear datos de censos
    parsed_censos = censos_loader()
    log.debug(parsed_censos)
    log.debug(dict_centros_de_acopio)

    # Algoritmo de distribución principal
    censo_spread(data_encuesta=parsed_data, data_censos=parsed_censos)

    log.info("Script finalizado")


###############################################################################
#                                INIT SCRIPT                                  #
###############################################################################
if __name__ == '__main__':
    args = setup()
    main()
